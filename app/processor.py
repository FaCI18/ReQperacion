"""Document text extraction processor for ReQperacion.

Extracts text content from various document types:
- PDF (.pdf) via PyMuPDF
- Word (.docx) via python-docx
- Excel (.xlsx) via openpyxl
- Plain text (.txt) via built-in
"""

import os
import logging

logger = logging.getLogger(__name__)


def extract_text_from_pdf(file_path: str) -> str:
    """Extract text from a PDF file using PyMuPDF."""
    try:
        import fitz  # PyMuPDF
        text_parts = []
        with fitz.open(file_path) as doc:
            for page_num, page in enumerate(doc):
                page_text = page.get_text()
                if page_text.strip():
                    text_parts.append(page_text)
        return "\n".join(text_parts)
    except ImportError:
        logger.error("PyMuPDF (fitz) is not installed.")
        return ""
    except Exception as e:
        logger.error(f"Error extracting text from PDF: {e}")
        return ""


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from a Word document."""
    try:
        from docx import Document
        doc = Document(file_path)
        text_parts = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_parts.append(paragraph.text)
        return "\n".join(text_parts)
    except ImportError:
        logger.error("python-docx is not installed.")
        return ""
    except Exception as e:
        logger.error(f"Error extracting text from DOCX: {e}")
        return ""


def extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from an Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(
                    str(cell) for cell in row if cell is not None
                )
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    except ImportError:
        logger.error("openpyxl is not installed.")
        return ""
    except Exception as e:
        logger.error(f"Error extracting text from XLSX: {e}")
        return ""


def extract_text_from_txt(file_path: str) -> str:
    """Extract text from a plain text file."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
    except Exception as e:
        logger.error(f"Error reading text file: {e}")
        return ""


# Mapping of file extensions to extraction functions
EXTRACTORS = {
    ".pdf": extract_text_from_pdf,
    ".docx": extract_text_from_docx,
    ".doc": extract_text_from_docx,
    ".xlsx": extract_text_from_xlsx,
    ".xls": extract_text_from_xlsx,
    ".txt": extract_text_from_txt,
    ".csv": extract_text_from_txt,
    ".md": extract_text_from_txt,
    ".json": extract_text_from_txt,
    ".xml": extract_text_from_txt,
    ".html": extract_text_from_txt,
    ".htm": extract_text_from_txt,
    ".py": extract_text_from_txt,
    ".js": extract_text_from_txt,
    ".ts": extract_text_from_txt,
    ".css": extract_text_from_txt,
    ".sql": extract_text_from_txt,
}


def extract_text(file_path: str) -> str:
    """
    Extract text from a document based on its file extension.
    Returns the extracted text or empty string if unsupported.
    """
    if not os.path.exists(file_path):
        logger.warning(f"File not found: {file_path}")
        return ""

    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    extractor = EXTRACTORS.get(ext)
    if extractor:
        logger.info(f"Extracting text from {file_path} using {extractor.__name__}")
        return extractor(file_path)

    logger.info(f"No extractor available for extension: {ext}")
    return ""
